#!/usr/bin/env python 
# -*- coding: utf-8 -*-
# @Time    : 2023/3/20 10:37
# @Author  : Jacque
# @Site    : 
# @File    : data.py
# @Software: PyCharm
# @Mail    : Jacquewong@stu.jluzh.edu.cn
# @Desc    :


from openpyxl import *
from openpyxl.utils import get_column_letter


def update_excel(process: list):
    title_row = ["step", "res", "lc", "event", "para"
                 # , "waiting time", "evade value", "remark"
                 ]
    wb = load_workbook("config/data.xlsx")
    ws = wb.active
    for index in process:
        for key in dict.keys(index):
            ws.cell(row=int(index["step"]) + 1, column=title_row.index(key) + 1, value=index[key])
    wb.save("config/data.xlsx")


def data_init():
    title_row = [
        "step",
        "resource",
        "loop count",
        "event",
        "parameter",
        "waiting time",
        "evade value",
        "remark"]
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.print_title_cols = 'A:A'  # the first col
    ws.print_title_rows = '1:1'  # the first row
    for col in range(0, len(title_row)):
        ws.cell(column=col + 1, row=1, value=title_row[col])
        ws.column_dimensions[get_column_letter(col + 1)].auto_size = True
    wb.save("./../config/data.xlsx")

# data_init()
